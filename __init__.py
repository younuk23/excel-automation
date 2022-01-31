from openpyxl import load_workbook
from openpyxl.styles import Font, Border, PatternFill, Side, Alignment, NamedStyle
from openpyxl.utils import get_column_letter
from copy import copy

wb = load_workbook('test.xlsx')

ws = wb.active
ws.title = "일일출고형식"

columns = {
    "CS": ws["A"],
    "company": ws["B"],
    "maker": ws["C"],
    "mall": ws["D"],
    "receiver": ws["E"],
    "phone_number":  ws["F"],
    "second_number":  ws["G"],
    "orderer": ws["H"],
    "order_number": ws["I"],
    "id": ws["J"],
    "delivery_number": ws["K"],
    "invoice_number": ws["L"],
    "delivery_status": ws["M"],
    "product_name": ws["N"],
    "option": ws["O"],
    "quantity": ws["P"],
    "box": ws["Q"],
    "regional": ws["R"],
    "address": ws["S"],
    "delivery_company": ws["T"],
    "prepaid": ws["U"],
    "pay_on_delivery": ws["V"],
    "product_cost": ws["W"],
    "delivery_man_cost": ws["X"],
    "order_enrollment_day": ws["Y"],
    "warehousing_day": ws["Z"],
    "designated_day": ws["AA"],
    "departure_from_branch_store_day": ws["AB"],
    "arrived_at_branch_store": ws["AC"],
    "delivery_start": ws["AD"],
    "delivery_complete": ws["AE"],
    "recovery": ws["AF"],
    "delivery_message": ws["AG"],
    "memo": ws["AH"],
    "head_office_memo": ws["AI"],
    "modified_man": ws["AJ"]
}


def merge_CS_to_company():
    for n, cell in enumerate(columns["CS"]):
        if n == 0:
            continue

        origin_value = columns["company"][n].value

        columns["company"][n].value = (
                f"{origin_value}({cell.value})"
                if cell.value is not None
                else origin_value
                )


def change_column_title_for_box():
    columns["box"][0].value = "수량"

def write_necessary_columns():
    ws.insert_cols(0, 11)

    necessary_columns = [
        columns["company"],
        columns["product_name"],
        columns["option"],
        columns["box"],
        columns["receiver"],
        columns["phone_number"],
        columns["second_number"],
        columns["address"],
        columns["pay_on_delivery"],
        columns["prepaid"],
        columns["delivery_message"]
        ]

    for col in ws.iter_cols(min_col=1, max_col=11):
        for cell in col:
            source = necessary_columns[cell.column - 1]
            source_cell = source[cell.row - 1]
            cell.value = source_cell.value
            cell.font = copy(source_cell.font)
            cell.fill = copy(source_cell.fill)
            cell.border = copy(source_cell.border)

    ws.delete_cols(12, 99)


def write_region_dict():
    pass_region = ["경남", "경북"]

    region_dict = {
            "regions": {},
            "rows": {}
            }

    for n, cell in enumerate(columns["address"]):
        if cell.value == "주소":
            continue
        splited = cell.value.split()
        region = splited[1] if splited[0] in pass_region else splited[0]

        if f"{region}_rows" in region_dict["rows"]:
            region_dict["rows"][f"{region}_rows"].append(n)
        else:
            region_dict["rows"][f"{region}_rows"] = [n]

        if region not in region_dict["regions"]:
            region_dict["regions"][region] = wb.create_sheet(region)

    return region_dict


def write_region_in_worksheet(worksheet, region):
    worksheet.merge_cells('A1:D1')
    title_cell = worksheet["A1"]
    title_cell.value = region
    title_cell.font = Font(size=26, bold=True)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = PatternFill("solid", fgColor="E0E0E1")

    title_border = Side(border_style="double", color="000000")
    for row in worksheet.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.border = Border(top=title_border, left=title_border, right=title_border, bottom=title_border)


def write_region_in_worksheets(region_dict):
    for key, value in region_dict.items():
        write_region_in_worksheet(value, key)


def write_values_in_region_sheet(region_dict):
    for region, worksheet in region_dict["regions"].items():
        worksheet.delete_cols(5)

        row_list = []
        for n, row_index in enumerate(region_dict["rows"][f"{region}_rows"]):
            row = {
                    "receiver":columns['receiver'][row_index].value,
                    "product_name":columns['product_name'][row_index].value,
                    "option":columns["option"][row_index].value,
                    "quantity":columns["box"][row_index].value
                    }
            row_list.append(row)
        row_list.sort(key=lambda row: row["receiver"])

        for n, row in enumerate(row_list):
            worksheet[f"A{n+2}"] = row["receiver"]
            worksheet[f"B{n+2}"] = row["product_name"]
            worksheet[f"C{n+2}"] = row["option"]
            worksheet[f"D{n+2}"] = row["quantity"]

        worksheet.delete_rows(len(row_list) + 2, 99)


def styling_region_sheet(region_sheets):
    region_style = NamedStyle(name="region_style")
    region_style.font = Font(size=18)
    region_style.alignment = Alignment(wrap_text=True)
    border_style = Side(border_style="thin", color="000000")
    region_style.border = Border(top=border_style, left=border_style, bottom=border_style, right=border_style)
    wb.add_named_style(region_style)

    for worksheet in region_sheets.values():
        for n, rows in enumerate(worksheet.iter_rows(min_row=1)):
            if n == 0:
                continue

            for cell in rows:
                cell.style = region_style

def adjust_column_width_for(region_sheets):
    widths = [30, 50, 50, 20]
    for worksheet in region_sheets.values():
        for n, width in enumerate(widths, 1):
            worksheet.column_dimensions[get_column_letter(n)].width = width

def init():
    merge_CS_to_company()
    change_column_title_for_box()
    write_necessary_columns()
    region_dict = write_region_dict()
    write_region_in_worksheets(region_dict["regions"])
    write_values_in_region_sheet(region_dict)
    styling_region_sheet(region_dict["regions"])
    adjust_column_width_for(region_dict["regions"])


init()


wb.save("test_result.xlsx")
wb.close()
